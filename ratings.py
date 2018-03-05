# ratings GUI
# by a-colabella & havoc17
from openpyxl import worksheet
from openpyxl import load_workbook

# Weights for duel/gf scores
duel_weight = .05
gf_weight = 1 - duel_weight

# a Player
class Player:
    username = ''
    duel = 0
    gf = 0
    w_av = 0
    pred_rank = 0
    real_rank = 0
    
    def __init__(self, username):
        self.username = username

    w_av = (duel_weight * duel) + (gf_weight * gf)


# Excel input
wb = load_workbook('Scores.xlsx')
ws = wb.active

# List of players
player_list = []

# A dictionary corresponding player names to player objects
player_dict = {}

# Retrieve duel scores
for count in range(2, 32):
    x = Player(str(ws.cell(count, 1).value))
    x.duel = int(ws.cell(count, 2).value)
    player_list.append(x)
    player_dict[x.username] = x

# Retrieve groupfighting scores
for count in range(2, 32):
    x = player_dict[str(ws.cell(count, 4).value)]
    x.gf = int(ws.cell(count, 5).value)
    print("Player: " + x.username + " Duel: " + str(x.duel) + " Gf: " + str(x.gf) + "\n")

    
# Retrieve predicted rank
